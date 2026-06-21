#!/usr/bin/env python3
"""Generate a sample STM (Source-to-Target Mapping) Excel workbook.

Produces a complete 8-sheet xlsx for the Patient 360 healthcare pipeline
using openpyxl. Run with:

    uv run python mapping-analyst-plugin/skills/create-stm/examples/sample-stm.py

Output: mapping-analyst-plugin/skills/create-stm/examples/sample-stm.xlsx
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Formatting constants
# ---------------------------------------------------------------------------

HEADER_FONT = Font(bold=True, size=11)
SOURCE_FILL = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
TARGET_FILL = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
TRANSFORM_FILL = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
META_FILL = PatternFill(start_color="F2F3F4", end_color="F2F3F4", fill_type="solid")

WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")

# SQL CASE expressions used in Bronze-to-Silver mappings (long by nature)
_GENDER_CASE = (
    "CASE UPPER(TRIM(gender)) WHEN 'M' THEN 'MALE' WHEN 'F' THEN 'FEMALE' ELSE 'UNKNOWN' END"
)
_RACE_CASE = (
    "CASE LOWER(TRIM(race)) WHEN 'white' THEN '2106-3'"
    " WHEN 'black' THEN '2054-5' WHEN 'asian' THEN '2028-9'"
    " ELSE 'UNK' END"
)
_ETHNICITY_CASE = (
    "CASE LOWER(TRIM(ethnicity)) WHEN 'hispanic' THEN '2135-2'"
    " WHEN 'nonhispanic' THEN '2186-5' ELSE 'UNK' END"
)
_ENCOUNTER_CASE = (
    "CASE LOWER(TRIM(encounterclass))"
    " WHEN 'ambulatory' THEN '371883000'"
    " WHEN 'emergency' THEN '50849002'"
    " WHEN 'inpatient' THEN '32485007'"
    " WHEN 'wellness' THEN '410620009'"
    " ELSE 'UNMAPPED' END"
)
_GENDER_FULL_CASE = (
    "CASE UPPER(TRIM(gender)) WHEN 'M' THEN 'MALE'"
    " WHEN 'F' THEN 'FEMALE' WHEN 'O' THEN 'OTHER'"
    " ELSE 'UNKNOWN' END"
)


def _add_headers(ws, headers, col_fills=None, widths=None):
    """Add formatted headers to row 1."""
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        if col_fills and col <= len(col_fills):
            cell.fill = col_fills[col - 1]
        else:
            cell.fill = META_FILL
    # Freeze top row
    ws.freeze_panes = "A2"
    # Column widths
    if widths:
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w
    else:
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 22
    # Auto-filter
    if len(headers) > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _add_rows(ws, rows, start_row=2):
    """Add data rows starting at the given row."""
    for r_idx, row in enumerate(rows, start_row):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = WRAP_ALIGN


def main():
    wb = openpyxl.Workbook()

    # -----------------------------------------------------------------------
    # Sheet 1: Summary
    # -----------------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.column_dimensions["A"].width = 25
    ws1.column_dimensions["B"].width = 60

    metadata = [
        ("Version", "1.0"),
        ("Created", "2026-03-16"),
        ("Author", "Mapping Analyst Agent"),
        ("Status", "Draft"),
        ("Pipeline", "Patient 360 Analytics"),
        ("DMS Ref", "DMS-2026-03-15-patient-360.md"),
        ("HLD Ref", "HLD-2026-03-14-pipeline.md"),
        ("DRD Ref", "DRD-2026-01-29-patient-360.md"),
        ("", ""),
        (
            "Approach",
            "Medallion: source → bronze (raw) → silver (cleansed) → gold (dimensional)",
        ),
        (
            "Bronze Strategy",
            "Column pass-through + metadata (ingestion_timestamp, source_file, batch_id)",
        ),
        (
            "Silver Strategy",
            "Type casting, null handling, dedup via ROW_NUMBER, SNOMED/RxNorm code standardization",
        ),
        (
            "Gold Strategy",
            "Star schema with SCD Type 2 for dim_patient, surrogate keys via IDENTITY sequence",
        ),
    ]

    for row_idx, (key, val) in enumerate(metadata, 1):
        k_cell = ws1.cell(row=row_idx, column=1, value=key)
        k_cell.font = Font(bold=True) if key else Font()
        ws1.cell(row=row_idx, column=2, value=val)

    # Traceability table starting at row 16
    ws1.cell(row=16, column=1, value="DMS Section").font = HEADER_FONT
    ws1.cell(row=16, column=2, value="STM Sheet").font = HEADER_FONT
    ws1.cell(row=16, column=3, value="Description").font = HEADER_FONT
    ws1.column_dimensions["C"].width = 40
    trace_rows = [
        ("DMS §2 - Bronze Layer Schemas", "Source-to-Bronze", "Column pass-through mappings"),
        ("DMS §3 - Silver Layer Schemas", "Bronze-to-Silver", "Cleansing and transformation rules"),
        ("DMS §4 - Gold Dimensions", "Silver-to-Gold", "Dimensional model mappings"),
        ("DMS §5 - Gold Facts", "Silver-to-Gold", "Fact table population"),
        ("DMS §6 - SCD Strategy", "Silver-to-Gold", "SCD Type 1/2 merge logic"),
        ("DMS §7 - Naming Conventions", "All sheets", "Column naming standards"),
    ]
    for r_idx, row in enumerate(trace_rows, 17):
        for c_idx, val in enumerate(row, 1):
            ws1.cell(row=r_idx, column=c_idx, value=val)

    # -----------------------------------------------------------------------
    # Sheet 2: Source-to-Bronze
    # -----------------------------------------------------------------------
    ws2 = wb.create_sheet("Source-to-Bronze")
    h2 = [
        "source_table",
        "source_column",
        "source_type",
        "target_table",
        "target_column",
        "target_type",
        "transformation",
        "notes",
    ]
    fills2 = [
        SOURCE_FILL,
        SOURCE_FILL,
        SOURCE_FILL,
        TARGET_FILL,
        TARGET_FILL,
        TARGET_FILL,
        TRANSFORM_FILL,
        META_FILL,
    ]
    _add_headers(ws2, h2, fills2)
    rows2 = [
        (
            "synthea.patients",
            "id",
            "VARCHAR",
            "brz_patients",
            "id",
            "VARCHAR",
            "PASS_THROUGH",
            "Primary key",
        ),
        (
            "synthea.patients",
            "birthdate",
            "VARCHAR",
            "brz_patients",
            "birthdate",
            "VARCHAR",
            "PASS_THROUGH",
            "Date as string",
        ),
        (
            "synthea.patients",
            "deathdate",
            "VARCHAR",
            "brz_patients",
            "deathdate",
            "VARCHAR",
            "PASS_THROUGH",
            "Nullable",
        ),
        (
            "synthea.patients",
            "ssn",
            "VARCHAR",
            "brz_patients",
            "ssn",
            "VARCHAR",
            "PASS_THROUGH",
            "PII — encrypt in silver",
        ),
        (
            "synthea.patients",
            "first",
            "VARCHAR",
            "brz_patients",
            "first",
            "VARCHAR",
            "PASS_THROUGH",
            "",
        ),
        (
            "synthea.patients",
            "last",
            "VARCHAR",
            "brz_patients",
            "last",
            "VARCHAR",
            "PASS_THROUGH",
            "",
        ),
        (
            "synthea.patients",
            "race",
            "VARCHAR",
            "brz_patients",
            "race",
            "VARCHAR",
            "PASS_THROUGH",
            "CDC code mapping in silver",
        ),
        (
            "synthea.patients",
            "ethnicity",
            "VARCHAR",
            "brz_patients",
            "ethnicity",
            "VARCHAR",
            "PASS_THROUGH",
            "CDC code mapping in silver",
        ),
        (
            "synthea.patients",
            "gender",
            "VARCHAR",
            "brz_patients",
            "gender",
            "VARCHAR",
            "PASS_THROUGH",
            "M/F standardization in silver",
        ),
        (
            "synthea.patients",
            "city",
            "VARCHAR",
            "brz_patients",
            "city",
            "VARCHAR",
            "PASS_THROUGH",
            "",
        ),
        (
            "synthea.patients",
            "state",
            "VARCHAR",
            "brz_patients",
            "state",
            "VARCHAR",
            "PASS_THROUGH",
            "",
        ),
        (
            "synthea.patients",
            "zip",
            "VARCHAR",
            "brz_patients",
            "zip",
            "VARCHAR",
            "PASS_THROUGH",
            "",
        ),
        (
            "synthea.patients",
            "(injected)",
            "TIMESTAMP",
            "brz_patients",
            "ingestion_timestamp",
            "TIMESTAMP",
            "CURRENT_TIMESTAMP",
            "Metadata injection",
        ),
        (
            "synthea.patients",
            "(injected)",
            "VARCHAR",
            "brz_patients",
            "source_file",
            "VARCHAR",
            "'synthea/patients.csv'",
            "Metadata injection",
        ),
        (
            "synthea.patients",
            "(injected)",
            "VARCHAR",
            "brz_patients",
            "batch_id",
            "VARCHAR",
            "'{run_date}_patients'",
            "Metadata injection",
        ),
        (
            "synthea.encounters",
            "Id",
            "VARCHAR",
            "brz_encounters",
            "id",
            "VARCHAR",
            "PASS_THROUGH",
            "Primary key",
        ),
        (
            "synthea.encounters",
            "PATIENT",
            "VARCHAR",
            "brz_encounters",
            "patient",
            "VARCHAR",
            "PASS_THROUGH",
            "FK to patients",
        ),
        (
            "synthea.encounters",
            "ENCOUNTERCLASS",
            "VARCHAR",
            "brz_encounters",
            "encounterclass",
            "VARCHAR",
            "PASS_THROUGH",
            "SNOMED lookup in silver",
        ),
        (
            "synthea.encounters",
            "START",
            "VARCHAR",
            "brz_encounters",
            "start",
            "VARCHAR",
            "PASS_THROUGH",
            "Timestamp string",
        ),
        (
            "synthea.encounters",
            "STOP",
            "VARCHAR",
            "brz_encounters",
            "stop",
            "VARCHAR",
            "PASS_THROUGH",
            "Timestamp string",
        ),
    ]
    _add_rows(ws2, rows2)

    # -----------------------------------------------------------------------
    # Sheet 3: Bronze-to-Silver
    # -----------------------------------------------------------------------
    ws3 = wb.create_sheet("Bronze-to-Silver")
    h3 = [
        "source_table",
        "source_column",
        "source_type",
        "target_table",
        "target_column",
        "target_type",
        "transformation",
        "null_handling",
        "default_value",
        "business_rule_ref",
        "dms_ref",
    ]
    fills3 = [
        SOURCE_FILL,
        SOURCE_FILL,
        SOURCE_FILL,
        TARGET_FILL,
        TARGET_FILL,
        TARGET_FILL,
        TRANSFORM_FILL,
        TRANSFORM_FILL,
        TRANSFORM_FILL,
        META_FILL,
        META_FILL,
    ]
    _add_headers(ws3, h3, fills3)
    rows3 = [
        (
            "brz_patients",
            "id",
            "VARCHAR",
            "slv_patients",
            "patient_id",
            "VARCHAR",
            "TRIM(id)",
            "REJECT",
            "",
            "BR-001",
            "DMS §3.1",
        ),
        (
            "brz_patients",
            "birthdate",
            "VARCHAR",
            "slv_patients",
            "date_of_birth",
            "DATE",
            "CAST(birthdate AS DATE)",
            "REJECT",
            "",
            "BR-002",
            "DMS §3.1",
        ),
        (
            "brz_patients",
            "deathdate",
            "VARCHAR",
            "slv_patients",
            "death_date",
            "DATE",
            "TRY_CAST(deathdate AS DATE)",
            "PASS NULL",
            "",
            "BR-003",
            "DMS §3.1",
        ),
        (
            "brz_patients",
            "first",
            "VARCHAR",
            "slv_patients",
            "first_name",
            "VARCHAR",
            "INITCAP(TRIM(first))",
            "DEFAULT",
            "UNKNOWN",
            "BR-004",
            "DMS §3.1",
        ),
        (
            "brz_patients",
            "last",
            "VARCHAR",
            "slv_patients",
            "last_name",
            "VARCHAR",
            "INITCAP(TRIM(last))",
            "DEFAULT",
            "UNKNOWN",
            "BR-005",
            "DMS §3.1",
        ),
        (
            "brz_patients",
            "gender",
            "VARCHAR",
            "slv_patients",
            "gender_code",
            "VARCHAR",
            _GENDER_CASE,
            "DEFAULT",
            "UNKNOWN",
            "BR-006",
            "DMS §3.1",
        ),
        (
            "brz_patients",
            "race",
            "VARCHAR",
            "slv_patients",
            "race_code",
            "VARCHAR",
            _RACE_CASE,
            "DEFAULT",
            "UNK",
            "BR-007",
            "DMS §3.1",
        ),
        (
            "brz_patients",
            "ethnicity",
            "VARCHAR",
            "slv_patients",
            "ethnicity_code",
            "VARCHAR",
            _ETHNICITY_CASE,
            "DEFAULT",
            "UNK",
            "BR-008",
            "DMS §3.1",
        ),
        (
            "brz_patients",
            "city",
            "VARCHAR",
            "slv_patients",
            "city",
            "VARCHAR",
            "TRIM(city)",
            "DEFAULT",
            "UNKNOWN",
            "BR-009",
            "DMS §3.1",
        ),
        (
            "brz_patients",
            "state",
            "VARCHAR",
            "slv_patients",
            "state",
            "VARCHAR",
            "UPPER(TRIM(state))",
            "DEFAULT",
            "UNKNOWN",
            "BR-010",
            "DMS §3.1",
        ),
        (
            "brz_patients",
            "zip",
            "VARCHAR",
            "slv_patients",
            "zip_code",
            "VARCHAR",
            "TRIM(zip)",
            "DEFAULT",
            "00000",
            "BR-011",
            "DMS §3.1",
        ),
        (
            "brz_patients",
            "(dedup)",
            "",
            "slv_patients",
            "(dedup)",
            "",
            "ROW_NUMBER() OVER (PARTITION BY id ORDER BY ingestion_timestamp DESC) = 1",
            "N/A",
            "",
            "BR-012",
            "DMS §3.1",
        ),
        (
            "brz_encounters",
            "id",
            "VARCHAR",
            "slv_encounters",
            "encounter_id",
            "VARCHAR",
            "TRIM(id)",
            "REJECT",
            "",
            "BR-020",
            "DMS §3.2",
        ),
        (
            "brz_encounters",
            "patient",
            "VARCHAR",
            "slv_encounters",
            "patient_id",
            "VARCHAR",
            "TRIM(patient)",
            "REJECT",
            "",
            "BR-021",
            "DMS §3.2",
        ),
        (
            "brz_encounters",
            "encounterclass",
            "VARCHAR",
            "slv_encounters",
            "encounter_type_code",
            "VARCHAR",
            _ENCOUNTER_CASE,
            "DEFAULT",
            "UNMAPPED",
            "BR-022",
            "DMS §3.2",
        ),
        (
            "brz_encounters",
            "start",
            "VARCHAR",
            "slv_encounters",
            "encounter_start",
            "TIMESTAMP",
            "STRPTIME(start, '%Y-%m-%dT%H:%M:%SZ')",
            "REJECT",
            "",
            "BR-023",
            "DMS §3.2",
        ),
        (
            "brz_encounters",
            "stop",
            "VARCHAR",
            "slv_encounters",
            "encounter_end",
            "TIMESTAMP",
            "TRY_CAST(stop AS TIMESTAMP)",
            "PASS NULL",
            "",
            "BR-024",
            "DMS §3.2",
        ),
    ]
    _add_rows(ws3, rows3)

    # -----------------------------------------------------------------------
    # Sheet 4: Silver-to-Gold
    # -----------------------------------------------------------------------
    ws4 = wb.create_sheet("Silver-to-Gold")
    h4 = [
        "target_table",
        "target_column",
        "target_type",
        "source_expression",
        "join_logic",
        "scd_type",
        "grain",
        "dms_ref",
    ]
    fills4 = [
        TARGET_FILL,
        TARGET_FILL,
        TARGET_FILL,
        TRANSFORM_FILL,
        TRANSFORM_FILL,
        META_FILL,
        META_FILL,
        META_FILL,
    ]
    _add_headers(ws4, h4, fills4)
    rows4 = [
        (
            "dim_patient",
            "patient_sk",
            "BIGINT",
            "GENERATED ALWAYS AS IDENTITY",
            "",
            "N/A",
            "one row per patient version",
            "DMS §5.1",
        ),
        (
            "dim_patient",
            "patient_id",
            "VARCHAR",
            "slv_patients.patient_id",
            "",
            "Type 2",
            "natural key",
            "DMS §5.1",
        ),
        (
            "dim_patient",
            "first_name",
            "VARCHAR",
            "slv_patients.first_name",
            "",
            "Type 1",
            "",
            "DMS §5.1",
        ),
        (
            "dim_patient",
            "last_name",
            "VARCHAR",
            "slv_patients.last_name",
            "",
            "Type 1",
            "",
            "DMS §5.1",
        ),
        (
            "dim_patient",
            "date_of_birth",
            "DATE",
            "slv_patients.date_of_birth",
            "",
            "Type 1",
            "",
            "DMS §5.1",
        ),
        (
            "dim_patient",
            "current_age",
            "INTEGER",
            "DATE_DIFF('year', slv_patients.date_of_birth, CURRENT_DATE)",
            "",
            "Type 1",
            "derived",
            "DMS §5.1",
        ),
        (
            "dim_patient",
            "is_deceased",
            "BOOLEAN",
            "slv_patients.death_date IS NOT NULL",
            "",
            "Type 1",
            "derived",
            "DMS §5.1",
        ),
        (
            "dim_patient",
            "gender_code",
            "VARCHAR",
            "slv_patients.gender_code",
            "",
            "Type 1",
            "",
            "DMS §5.1",
        ),
        (
            "dim_patient",
            "city",
            "VARCHAR",
            "slv_patients.city",
            "",
            "Type 2",
            "tracked for address history",
            "DMS §5.1",
        ),
        (
            "dim_patient",
            "state",
            "VARCHAR",
            "slv_patients.state",
            "",
            "Type 2",
            "tracked for address history",
            "DMS §5.1",
        ),
        (
            "dim_patient",
            "zip_code",
            "VARCHAR",
            "slv_patients.zip_code",
            "",
            "Type 2",
            "tracked for address history",
            "DMS §5.1",
        ),
        (
            "dim_patient",
            "effective_from",
            "DATE",
            "CURRENT_DATE (on insert/update)",
            "",
            "Type 2",
            "SCD2 bookkeeping",
            "DMS §6.1",
        ),
        (
            "dim_patient",
            "effective_to",
            "DATE",
            "'9999-12-31' (current) or expire date",
            "",
            "Type 2",
            "SCD2 bookkeeping",
            "DMS §6.1",
        ),
        (
            "dim_patient",
            "is_current",
            "BOOLEAN",
            "TRUE (on insert), FALSE (on expire)",
            "",
            "Type 2",
            "SCD2 bookkeeping",
            "DMS §6.1",
        ),
        (
            "fact_encounter",
            "encounter_sk",
            "BIGINT",
            "GENERATED ALWAYS AS IDENTITY",
            "",
            "N/A",
            "one row per encounter",
            "DMS §5.2",
        ),
        (
            "fact_encounter",
            "encounter_id",
            "VARCHAR",
            "slv_encounters.encounter_id",
            "",
            "N/A",
            "degenerate dimension",
            "DMS §5.2",
        ),
        (
            "fact_encounter",
            "patient_sk",
            "BIGINT",
            "dim_patient.patient_sk",
            "JOIN dim_patient ON patient_id AND is_current = TRUE",
            "N/A",
            "FK lookup",
            "DMS §5.2",
        ),
        (
            "fact_encounter",
            "encounter_type_code",
            "VARCHAR",
            "slv_encounters.encounter_type_code",
            "",
            "N/A",
            "",
            "DMS §5.2",
        ),
        (
            "fact_encounter",
            "encounter_start",
            "TIMESTAMP",
            "slv_encounters.encounter_start",
            "",
            "N/A",
            "",
            "DMS §5.2",
        ),
        (
            "fact_encounter",
            "encounter_end",
            "TIMESTAMP",
            "slv_encounters.encounter_end",
            "",
            "N/A",
            "",
            "DMS §5.2",
        ),
    ]
    _add_rows(ws4, rows4)

    # -----------------------------------------------------------------------
    # Sheet 5: Code Systems
    # -----------------------------------------------------------------------
    ws5 = wb.create_sheet("Code Systems")
    h5 = ["code_system", "source_value_pattern", "target_value", "case_expression", "notes"]
    fills5 = [META_FILL, SOURCE_FILL, TARGET_FILL, TRANSFORM_FILL, META_FILL]
    _add_headers(ws5, h5, fills5, widths=[18, 22, 18, 50, 30])
    rows5 = [
        (
            "HL7 Gender",
            "M",
            "MALE",
            _GENDER_FULL_CASE,
            "HL7 AdministrativeGender",
        ),
        ("HL7 Gender", "F", "FEMALE", "", ""),
        ("HL7 Gender", "O", "OTHER", "", ""),
        (
            "SNOMED-CT",
            "ambulatory",
            "371883000",
            "CASE LOWER(encounterclass) WHEN 'ambulatory' THEN '371883000' ...",
            "Outpatient procedure",
        ),
        ("SNOMED-CT", "emergency", "50849002", "", "Emergency room admission"),
        ("SNOMED-CT", "inpatient", "32485007", "", "Hospital admission"),
        ("SNOMED-CT", "wellness", "410620009", "", "Well child visit"),
        ("SNOMED-CT", "urgentcare", "702927004", "", "Urgent care clinic"),
        (
            "CDC Race",
            "white",
            "2106-3",
            "CASE LOWER(race) WHEN 'white' THEN '2106-3' ...",
            "CDC Race & Ethnicity",
        ),
        ("CDC Race", "black", "2054-5", "", ""),
        ("CDC Race", "asian", "2028-9", "", ""),
        (
            "CDC Ethnicity",
            "hispanic",
            "2135-2",
            "CASE LOWER(ethnicity) WHEN 'hispanic' THEN '2135-2' ...",
            "CDC Race & Ethnicity",
        ),
        (
            "RxNorm",
            "Metformin 500mg",
            "860975",
            "Carried forward from source CODE column",
            "Validated against RxNorm",
        ),
        ("LOINC", "Body Height", "8302-2", "Carried forward from source CODE column", "Unit: cm"),
        ("LOINC", "Body Weight", "29463-7", "", "Unit: kg"),
        ("LOINC", "BMI", "39156-5", "", "Unit: kg/m2"),
    ]
    _add_rows(ws5, rows5)

    # -----------------------------------------------------------------------
    # Sheet 6: Null Handling
    # -----------------------------------------------------------------------
    ws6 = wb.create_sheet("Null Handling")
    h6 = [
        "table",
        "column",
        "layer",
        "criticality",
        "null_rate_observed",
        "action",
        "default_value",
        "business_rule_ref",
    ]
    fills6 = [
        TARGET_FILL,
        TARGET_FILL,
        META_FILL,
        TRANSFORM_FILL,
        META_FILL,
        TRANSFORM_FILL,
        TRANSFORM_FILL,
        META_FILL,
    ]
    _add_headers(ws6, h6, fills6)
    rows6 = [
        ("slv_patients", "patient_id", "silver", "HIGH", "0%", "REJECT", "", "Null PK not allowed"),
        (
            "slv_patients",
            "date_of_birth",
            "silver",
            "HIGH",
            "0%",
            "REJECT",
            "",
            "Required for age calc",
        ),
        (
            "slv_patients",
            "death_date",
            "silver",
            "LOW",
            "85%",
            "PASS NULL",
            "",
            "Most patients alive",
        ),
        ("slv_patients", "first_name", "silver", "MEDIUM", "0.1%", "DEFAULT", "UNKNOWN", ""),
        ("slv_patients", "last_name", "silver", "MEDIUM", "0.1%", "DEFAULT", "UNKNOWN", ""),
        ("slv_patients", "gender_code", "silver", "MEDIUM", "0.2%", "DEFAULT", "UNKNOWN", ""),
        (
            "slv_patients",
            "city",
            "silver",
            "MEDIUM",
            "1.5%",
            "DEFAULT",
            "UNKNOWN",
            "Address tracking",
        ),
        (
            "slv_patients",
            "state",
            "silver",
            "MEDIUM",
            "1.5%",
            "DEFAULT",
            "UNKNOWN",
            "Address tracking",
        ),
        ("slv_patients", "zip_code", "silver", "MEDIUM", "2%", "DEFAULT", "00000", ""),
        (
            "slv_encounters",
            "encounter_id",
            "silver",
            "HIGH",
            "0%",
            "REJECT",
            "",
            "Null PK not allowed",
        ),
        ("slv_encounters", "patient_id", "silver", "HIGH", "0%", "REJECT", "", "Required FK"),
        (
            "slv_encounters",
            "encounter_start",
            "silver",
            "HIGH",
            "0%",
            "REJECT",
            "",
            "Required timestamp",
        ),
        (
            "slv_encounters",
            "encounter_end",
            "silver",
            "LOW",
            "5%",
            "PASS NULL",
            "",
            "In-progress encounters",
        ),
    ]
    _add_rows(ws6, rows6)

    # -----------------------------------------------------------------------
    # Sheet 7: Edge Cases
    # -----------------------------------------------------------------------
    ws7 = wb.create_sheet("Edge Cases")
    h7 = ["category", "scenario", "affected_tables", "handling_rule", "severity", "dq_rule_ref"]
    fills7 = [META_FILL, TRANSFORM_FILL, TARGET_FILL, TRANSFORM_FILL, META_FILL, META_FILL]
    _add_headers(ws7, h7, fills7, widths=[20, 35, 25, 40, 12, 15])
    rows7 = [
        (
            "Orphan Records",
            "Encounter without matching patient in dim_patient",
            "fact_encounter",
            "Log to quarantine table, set patient_sk = -1 (unknown member)",
            "HIGH",
            "DQ-001",
        ),
        (
            "Future Dates",
            "birthdate > CURRENT_DATE",
            "slv_patients",
            "Reject record, log to DQ exceptions",
            "HIGH",
            "DQ-002",
        ),
        (
            "Future Dates",
            "encounter_start > CURRENT_TIMESTAMP",
            "slv_encounters",
            "Flag for review, allow with warning",
            "MEDIUM",
            "DQ-003",
        ),
        (
            "Duplicate Keys",
            "Multiple records with same patient id in bronze",
            "brz_patients → slv_patients",
            "Dedup via ROW_NUMBER, keep latest by ingestion_timestamp",
            "MEDIUM",
            "DQ-004",
        ),
        (
            "Schema Evolution",
            "New column added to source table",
            "brz_* tables",
            "Bronze layer passes all columns; silver mapping updated in next STM version",
            "LOW",
            "",
        ),
        (
            "Type Overflow",
            "VARCHAR date value unparseable",
            "brz_patients → slv_patients",
            "TRY_CAST returns NULL, null handling rules apply",
            "MEDIUM",
            "DQ-005",
        ),
        (
            "FK Integrity",
            "Silver patient_id not found in dim_patient",
            "fact_encounter",
            "Use unknown member (patient_sk = -1), log to DQ exceptions",
            "HIGH",
            "DQ-006",
        ),
        (
            "Empty Source",
            "Source table has 0 rows",
            "All bronze tables",
            "Skip load, log warning, do NOT truncate target",
            "HIGH",
            "DQ-007",
        ),
    ]
    _add_rows(ws7, rows7)

    # -----------------------------------------------------------------------
    # Sheet 8: Lineage
    # -----------------------------------------------------------------------
    ws8 = wb.create_sheet("Lineage")
    h8 = [
        "gold_table",
        "gold_column",
        "silver_expression",
        "bronze_column",
        "source_column",
        "transformation_chain",
    ]
    fills8 = [TARGET_FILL, TARGET_FILL, TRANSFORM_FILL, SOURCE_FILL, SOURCE_FILL, TRANSFORM_FILL]
    _add_headers(ws8, h8, fills8, widths=[18, 18, 35, 18, 22, 50])
    rows8 = [
        (
            "dim_patient",
            "patient_id",
            "slv_patients.patient_id",
            "brz_patients.id",
            "synthea.patients.id",
            "PASS_THROUGH → TRIM(id) → natural key",
        ),
        (
            "dim_patient",
            "first_name",
            "slv_patients.first_name",
            "brz_patients.first",
            "synthea.patients.first",
            "PASS_THROUGH → INITCAP(TRIM(first))",
        ),
        (
            "dim_patient",
            "last_name",
            "slv_patients.last_name",
            "brz_patients.last",
            "synthea.patients.last",
            "PASS_THROUGH → INITCAP(TRIM(last))",
        ),
        (
            "dim_patient",
            "date_of_birth",
            "slv_patients.date_of_birth",
            "brz_patients.birthdate",
            "synthea.patients.birthdate",
            "PASS_THROUGH → CAST AS DATE",
        ),
        (
            "dim_patient",
            "current_age",
            "DATE_DIFF('year', date_of_birth, CURRENT_DATE)",
            "brz_patients.birthdate",
            "synthea.patients.birthdate",
            "PASS_THROUGH → CAST AS DATE → DATE_DIFF",
        ),
        (
            "dim_patient",
            "is_deceased",
            "slv_patients.death_date IS NOT NULL",
            "brz_patients.deathdate",
            "synthea.patients.deathdate",
            "PASS_THROUGH → TRY_CAST AS DATE → IS NOT NULL",
        ),
        (
            "dim_patient",
            "gender_code",
            "slv_patients.gender_code",
            "brz_patients.gender",
            "synthea.patients.gender",
            "PASS_THROUGH → CASE WHEN standardization",
        ),
        (
            "dim_patient",
            "city",
            "slv_patients.city",
            "brz_patients.city",
            "synthea.patients.city",
            "PASS_THROUGH → TRIM → SCD2 tracked",
        ),
        (
            "fact_encounter",
            "encounter_id",
            "slv_encounters.encounter_id",
            "brz_encounters.id",
            "synthea.encounters.Id",
            "PASS_THROUGH → TRIM",
        ),
        (
            "fact_encounter",
            "patient_sk",
            "dim_patient.patient_sk (FK lookup)",
            "brz_encounters.patient",
            "synthea.encounters.PATIENT",
            "PASS_THROUGH → TRIM → JOIN dim_patient ON patient_id",
        ),
        (
            "fact_encounter",
            "encounter_type_code",
            "slv_encounters.encounter_type_code",
            "brz_encounters.encounterclass",
            "synthea.encounters.ENCOUNTERCLASS",
            "PASS_THROUGH → SNOMED-CT CASE mapping",
        ),
        (
            "fact_encounter",
            "encounter_start",
            "slv_encounters.encounter_start",
            "brz_encounters.start",
            "synthea.encounters.START",
            "PASS_THROUGH → STRPTIME",
        ),
    ]
    _add_rows(ws8, rows8)

    # -----------------------------------------------------------------------
    # Save
    # -----------------------------------------------------------------------
    output_path = Path(__file__).parent / "sample-stm.xlsx"
    wb.save(str(output_path))
    print(f"Sample STM generated: {output_path}")


if __name__ == "__main__":
    main()
