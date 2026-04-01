#!/usr/bin/env python3
"""STM (Source-to-Target Mapping) xlsx validator.

Validates STM Excel workbooks for completeness and quality.
Uses openpyxl to inspect sheet structure, column headers, data rows,
traceability, null handling, lineage, and formatting.

Usage:
    python validate_stm.py <stm-file.xlsx>
    python validate_stm.py --all <directory>
    python validate_stm.py --all <directory> --format json
"""

from __future__ import annotations

import argparse
import json
import sys
from dataclasses import dataclass, field
from enum import IntEnum
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

REQUIRED_SHEETS = [
    "Summary",
    "Source-to-Bronze",
    "Bronze-to-Silver",
    "Silver-to-Gold",
    "Code Systems",
    "Null Handling",
    "Edge Cases",
    "Lineage",
]

METADATA_KEYS = ["Version", "Created", "Author", "Status", "DMS Ref", "HLD Ref"]

SOURCE_TO_BRONZE_HEADERS = [
    "source_table",
    "source_column",
    "source_type",
    "target_table",
    "target_column",
    "target_type",
    "transformation",
    "notes",
]

BRONZE_TO_SILVER_HEADERS = [
    "source_table",
    "source_column",
    "source_type",
    "target_table",
    "target_column",
    "target_type",
    "transformation",
    "null_handling",
    "default_value",
    "business_rule_ref",
    "dms_ref",
]

SILVER_TO_GOLD_HEADERS = [
    "target_table",
    "target_column",
    "target_type",
    "source_expression",
    "join_logic",
    "scd_type",
    "grain",
    "dms_ref",
]

CODE_SYSTEMS_HEADERS = [
    "code_system",
    "source_value_pattern",
    "target_value",
    "case_expression",
    "notes",
]

NULL_HANDLING_HEADERS = [
    "table",
    "column",
    "layer",
    "criticality",
    "null_rate_observed",
    "action",
    "default_value",
    "business_rule_ref",
]

EDGE_CASES_HEADERS = [
    "category",
    "scenario",
    "affected_tables",
    "handling_rule",
    "severity",
    "dq_rule_ref",
]

LINEAGE_HEADERS = [
    "gold_table",
    "gold_column",
    "silver_expression",
    "bronze_column",
    "source_column",
    "transformation_chain",
]

SHEET_HEADERS = {
    "Source-to-Bronze": SOURCE_TO_BRONZE_HEADERS,
    "Bronze-to-Silver": BRONZE_TO_SILVER_HEADERS,
    "Silver-to-Gold": SILVER_TO_GOLD_HEADERS,
    "Code Systems": CODE_SYSTEMS_HEADERS,
    "Null Handling": NULL_HANDLING_HEADERS,
    "Edge Cases": EDGE_CASES_HEADERS,
    "Lineage": LINEAGE_HEADERS,
}

PLACEHOLDER_PATTERNS = ["TBD", "TODO", "PLACEHOLDER", "FIXME"]


# ---------------------------------------------------------------------------
# Validation framework
# ---------------------------------------------------------------------------


class ValidationLevel(IntEnum):
    """Severity levels for validation results."""

    INFO = 0
    WARNING = 1
    CRITICAL = 2


@dataclass
class ValidationResult:
    """A single validation finding."""

    level: ValidationLevel
    check_name: str
    message: str
    suggestion: str = ""


@dataclass
class ValidationReport:
    """Collects validation results for a single file."""

    file_path: str
    results: list[ValidationResult] = field(default_factory=list)

    def add_result(self, result: ValidationResult) -> None:
        self.results.append(result)

    @property
    def summary(self) -> dict[str, int]:
        counts: dict[str, int] = {"CRITICAL": 0, "WARNING": 0, "INFO": 0}
        for r in self.results:
            counts[r.level.name] += 1
        return counts

    def to_dict(self) -> dict:
        return {
            "file": self.file_path,
            "summary": self.summary,
            "results": [
                {
                    "level": r.level.name,
                    "check": r.check_name,
                    "message": r.message,
                    "suggestion": r.suggestion,
                }
                for r in self.results
            ],
        }

    def print_report(self) -> None:
        s = self.summary
        print(f"\n{'=' * 60}")
        print(f"STM Validation: {Path(self.file_path).name}")
        print(f"{'=' * 60}")
        print(f"  CRITICAL: {s['CRITICAL']}  WARNING: {s['WARNING']}  INFO: {s['INFO']}")
        print(f"{'-' * 60}")

        for level_name in ("CRITICAL", "WARNING", "INFO"):
            items = [r for r in self.results if r.level.name == level_name]
            if items:
                print(f"\n  [{level_name}]")
                for r in items:
                    print(f"    - [{r.check_name}] {r.message}")
                    if r.suggestion:
                        print(f"      Suggestion: {r.suggestion}")

        if s["CRITICAL"] == 0 and s["WARNING"] == 0:
            print("\n  PASS — No critical or warning issues found.")
        print()


# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------


def _get_headers(ws) -> list[str]:
    """Get lowercase header values from row 1 of a worksheet."""
    headers = []
    for cell in ws[1]:
        val = cell.value
        if val is not None:
            headers.append(str(val).strip().lower())
        else:
            headers.append("")
    return headers


def _data_row_count(ws) -> int:
    """Count non-empty data rows (rows after header)."""
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None and str(v).strip() != "" for v in row):
            count += 1
    return count


def _col_index(headers: list[str], name: str) -> int | None:
    """Find 0-based index of a column name (case-insensitive)."""
    name_lower = name.lower()
    for i, h in enumerate(headers):
        if h == name_lower:
            return i
    return None


# ---------------------------------------------------------------------------
# CRITICAL checks
# ---------------------------------------------------------------------------


def check_required_sheets(wb, report: ValidationReport) -> None:
    """All 8 required sheets must be present."""
    for sheet_name in REQUIRED_SHEETS:
        if sheet_name not in wb.sheetnames:
            report.add_result(
                ValidationResult(
                    level=ValidationLevel.CRITICAL,
                    check_name="required_sheets",
                    message=f"Missing required sheet: '{sheet_name}'",
                    suggestion=f"Add a sheet named '{sheet_name}' to the workbook",
                )
            )


VALID_STATUSES = {"Draft", "Updated - Pending Review", "Approved"}


def check_metadata(wb, report: ValidationReport) -> None:
    """Summary sheet must have required metadata keys."""
    if "Summary" not in wb.sheetnames:
        return
    ws = wb["Summary"]
    found_keys: dict[str, str | None] = {}
    for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        if row[0] is not None:
            key = str(row[0]).strip()
            found_keys[key] = str(row[1]).strip() if row[1] is not None else None

    for key in METADATA_KEYS:
        if key not in found_keys:
            report.add_result(
                ValidationResult(
                    level=ValidationLevel.CRITICAL,
                    check_name="metadata",
                    message=f"Missing metadata key in Summary: '{key}'",
                    suggestion=f"Add '{key}' to column A of the Summary sheet",
                )
            )

    # Validate Status field value
    status_value = found_keys.get("Status")
    if status_value and status_value not in VALID_STATUSES:
        report.add_result(
            ValidationResult(
                level=ValidationLevel.WARNING,
                check_name="metadata",
                message=f"Status has unrecognized value: '{status_value}'",
                suggestion=("Status must be one of: " + ", ".join(sorted(VALID_STATUSES))),
            )
        )


def check_source_to_bronze(wb, report: ValidationReport) -> None:
    """Source-to-Bronze sheet must have correct headers and data."""
    if "Source-to-Bronze" not in wb.sheetnames:
        return
    ws = wb["Source-to-Bronze"]
    headers = _get_headers(ws)

    required = ["source_table", "source_column", "target_table", "target_column", "transformation"]
    for h in required:
        if h not in headers:
            report.add_result(
                ValidationResult(
                    level=ValidationLevel.CRITICAL,
                    check_name="source_to_bronze",
                    message=f"Missing required header '{h}' in Source-to-Bronze",
                    suggestion=f"Add '{h}' column to the Source-to-Bronze sheet",
                )
            )
            return

    if _data_row_count(ws) < 1:
        report.add_result(
            ValidationResult(
                level=ValidationLevel.CRITICAL,
                check_name="source_to_bronze",
                message="Source-to-Bronze has no data rows",
                suggestion="Add at least one source-to-bronze mapping row",
            )
        )


def check_bronze_to_silver(wb, report: ValidationReport) -> None:
    """Bronze-to-Silver sheet must have correct headers and transformation data."""
    if "Bronze-to-Silver" not in wb.sheetnames:
        return
    ws = wb["Bronze-to-Silver"]
    headers = _get_headers(ws)

    required = ["source_table", "source_column", "target_table", "target_column", "transformation"]
    for h in required:
        if h not in headers:
            report.add_result(
                ValidationResult(
                    level=ValidationLevel.CRITICAL,
                    check_name="bronze_to_silver",
                    message=f"Missing required header '{h}' in Bronze-to-Silver",
                    suggestion=f"Add '{h}' column to the Bronze-to-Silver sheet",
                )
            )
            return

    if _data_row_count(ws) < 1:
        report.add_result(
            ValidationResult(
                level=ValidationLevel.CRITICAL,
                check_name="bronze_to_silver",
                message="Bronze-to-Silver has no data rows",
                suggestion="Add at least one bronze-to-silver transformation row",
            )
        )
        return

    # Check that transformation column has real values
    trans_idx = _col_index(headers, "transformation")
    if trans_idx is not None:
        has_real = False
        for row in ws.iter_rows(min_row=2, values_only=True):
            val = row[trans_idx] if trans_idx < len(row) else None
            if val and str(val).strip().upper() not in ("", "TBD", "TODO", "N/A"):
                has_real = True
                break
        if not has_real:
            report.add_result(
                ValidationResult(
                    level=ValidationLevel.CRITICAL,
                    check_name="bronze_to_silver",
                    message="Bronze-to-Silver has no real transformation expressions",
                    suggestion="Specify actual transformation logic (e.g., CAST, CASE, TRIM)",
                )
            )


def check_silver_to_gold(wb, report: ValidationReport) -> None:
    """Silver-to-Gold sheet must have correct headers and data."""
    if "Silver-to-Gold" not in wb.sheetnames:
        return
    ws = wb["Silver-to-Gold"]
    headers = _get_headers(ws)

    required = ["target_table", "target_column", "target_type", "source_expression"]
    for h in required:
        if h not in headers:
            report.add_result(
                ValidationResult(
                    level=ValidationLevel.CRITICAL,
                    check_name="silver_to_gold",
                    message=f"Missing required header '{h}' in Silver-to-Gold",
                    suggestion=f"Add '{h}' column to the Silver-to-Gold sheet",
                )
            )
            return

    if _data_row_count(ws) < 1:
        report.add_result(
            ValidationResult(
                level=ValidationLevel.CRITICAL,
                check_name="silver_to_gold",
                message="Silver-to-Gold has no data rows",
                suggestion="Add at least one silver-to-gold mapping row",
            )
        )


# ---------------------------------------------------------------------------
# WARNING checks
# ---------------------------------------------------------------------------


def check_dms_traceability(wb, report: ValidationReport) -> None:
    """DMS references should be populated in >=50% of mapping rows."""
    total = 0
    populated = 0

    for sheet_name in ("Bronze-to-Silver", "Silver-to-Gold"):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = _get_headers(ws)
        dms_idx = _col_index(headers, "dms_ref")
        if dms_idx is None:
            continue

        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(v is not None and str(v).strip() for v in row):
                total += 1
                val = row[dms_idx] if dms_idx < len(row) else None
                if val and str(val).strip() and str(val).strip().upper() not in ("TBD", "TODO", ""):
                    populated += 1

    if total > 0 and populated / total < 0.5:
        report.add_result(
            ValidationResult(
                level=ValidationLevel.WARNING,
                check_name="dms_traceability",
                message=(
                    f"DMS references populated in only {populated}/{total} rows "
                    f"({populated * 100 // total}%)"
                ),
                suggestion="Add [DMS §X.Y] references to dms_ref column for traceability",
            )
        )


def check_null_handling(wb, report: ValidationReport) -> None:
    """Null Handling sheet should have rows with criticality values."""
    if "Null Handling" not in wb.sheetnames:
        return
    ws = wb["Null Handling"]
    headers = _get_headers(ws)
    crit_idx = _col_index(headers, "criticality")

    has_criticality = False
    if crit_idx is not None:
        for row in ws.iter_rows(min_row=2, values_only=True):
            val = row[crit_idx] if crit_idx < len(row) else None
            if val and str(val).strip().upper() in ("HIGH", "MEDIUM", "LOW"):
                has_criticality = True
                break

    if not has_criticality:
        report.add_result(
            ValidationResult(
                level=ValidationLevel.WARNING,
                check_name="null_handling",
                message="Null Handling sheet has no rows with criticality values (HIGH/MEDIUM/LOW)",
                suggestion="Document null handling criticality for key fields",
            )
        )


def check_lineage(wb, report: ValidationReport) -> None:
    """Lineage sheet should have at least one complete lineage chain."""
    if "Lineage" not in wb.sheetnames:
        return
    ws = wb["Lineage"]
    headers = _get_headers(ws)

    required_cols = [
        "gold_table",
        "gold_column",
        "silver_expression",
        "bronze_column",
        "source_column",
        "transformation_chain",
    ]
    col_indices = [_col_index(headers, c) for c in required_cols]

    if any(idx is None for idx in col_indices):
        report.add_result(
            ValidationResult(
                level=ValidationLevel.WARNING,
                check_name="lineage",
                message="Lineage sheet missing expected column headers",
                suggestion=f"Expected columns: {', '.join(required_cols)}",
            )
        )
        return

    has_complete = False
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(
            idx < len(row) and row[idx] is not None and str(row[idx]).strip() for idx in col_indices
        ):
            has_complete = True
            break

    if not has_complete:
        report.add_result(
            ValidationResult(
                level=ValidationLevel.WARNING,
                check_name="lineage",
                message="No complete lineage chain found (all columns populated)",
                suggestion="Add at least one row with full gold→silver→bronze→source trace",
            )
        )


def check_code_system_mappings(wb, report: ValidationReport) -> None:
    """Code Systems sheet should have SNOMED/RxNorm/LOINC entries."""
    if "Code Systems" not in wb.sheetnames:
        return
    ws = wb["Code Systems"]
    headers = _get_headers(ws)
    cs_idx = _col_index(headers, "code_system")

    known_systems = {"SNOMED", "SNOMED-CT", "RXNORM", "LOINC"}
    has_known = False

    if cs_idx is not None:
        for row in ws.iter_rows(min_row=2, values_only=True):
            val = row[cs_idx] if cs_idx < len(row) else None
            if val and str(val).strip().upper().replace("-", "") in {
                s.replace("-", "") for s in known_systems
            }:
                has_known = True
                break

    if not has_known:
        report.add_result(
            ValidationResult(
                level=ValidationLevel.WARNING,
                check_name="code_system_mappings",
                message="Code Systems sheet has no SNOMED-CT, RxNorm, or LOINC entries",
                suggestion="Add healthcare code system mappings for standardization",
            )
        )


def check_edge_cases(wb, report: ValidationReport) -> None:
    """Edge Cases sheet should have at least one documented edge case."""
    if "Edge Cases" not in wb.sheetnames:
        return
    ws = wb["Edge Cases"]

    if _data_row_count(ws) < 1:
        report.add_result(
            ValidationResult(
                level=ValidationLevel.WARNING,
                check_name="edge_cases",
                message="Edge Cases sheet has no documented edge cases",
                suggestion=(
                    "Document edge cases like orphan records, future dates, schema evolution"
                ),
            )
        )


def check_column_headers(wb, report: ValidationReport) -> None:
    """Each mapping sheet should have expected standard column headers."""
    for sheet_name, expected_headers in SHEET_HEADERS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = _get_headers(ws)
        missing = [h for h in expected_headers if h.lower() not in headers]
        if missing:
            report.add_result(
                ValidationResult(
                    level=ValidationLevel.WARNING,
                    check_name="column_headers",
                    message=f"Sheet '{sheet_name}' missing standard headers: {', '.join(missing)}",
                    suggestion=f"Add columns: {', '.join(missing)}",
                )
            )


def check_transformation_expressions(wb, report: ValidationReport) -> None:
    """Transformation columns should not be blank or generic in most rows."""
    generic = {"TBD", "TODO", "N/A", "PLACEHOLDER", ""}

    for sheet_name in ("Source-to-Bronze", "Bronze-to-Silver"):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = _get_headers(ws)
        trans_idx = _col_index(headers, "transformation")
        if trans_idx is None:
            continue

        total = 0
        real = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(v is not None and str(v).strip() for v in row):
                total += 1
                val = row[trans_idx] if trans_idx < len(row) else None
                if val and str(val).strip().upper() not in generic:
                    real += 1

        if total > 0 and real / total < 0.7:
            report.add_result(
                ValidationResult(
                    level=ValidationLevel.WARNING,
                    check_name="transformation_expressions",
                    message=(
                        f"Sheet '{sheet_name}': only {real}/{total} rows have "
                        f"real transformation expressions"
                    ),
                    suggestion="Replace TBD/TODO with actual transformation logic",
                )
            )


# ---------------------------------------------------------------------------
# INFO checks
# ---------------------------------------------------------------------------


def check_placeholders(wb, report: ValidationReport) -> None:
    """Flag cells containing TBD/TODO/PLACEHOLDER text."""
    total_placeholders = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        count = 0
        for row in ws.iter_rows(values_only=True):
            for cell_val in row:
                if cell_val and any(p in str(cell_val).upper() for p in PLACEHOLDER_PATTERNS):
                    count += 1
        if count > 0:
            total_placeholders += count

    if total_placeholders > 0:
        report.add_result(
            ValidationResult(
                level=ValidationLevel.INFO,
                check_name="placeholders",
                message=(
                    f"Found {total_placeholders} placeholder(s) "
                    f"(TBD/TODO/PLACEHOLDER) across sheets"
                ),
                suggestion="Replace placeholders with actual values before finalizing",
            )
        )


def check_dq_references(wb, report: ValidationReport) -> None:
    """Check if DQ rule IDs are referenced in edge cases or null handling."""
    has_dq = False
    for sheet_name in ("Edge Cases", "Null Handling"):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            for cell_val in row:
                if cell_val and "DQ" in str(cell_val).upper():
                    has_dq = True
                    break
            if has_dq:
                break
        if has_dq:
            break

    if not has_dq:
        report.add_result(
            ValidationResult(
                level=ValidationLevel.INFO,
                check_name="dq_references",
                message="No DQ rule references found in Edge Cases or Null Handling sheets",
                suggestion="Consider referencing DQ rules (e.g., DQ-001) for traceability",
            )
        )


def check_formatting(wb, report: ValidationReport) -> None:
    """Check if headers are bold and sheets have auto-filter enabled."""
    issues = []
    for sheet_name in SHEET_HEADERS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        # Check bold headers
        if ws.max_row and ws.max_row >= 1:
            first_cell = ws.cell(row=1, column=1)
            if first_cell.value and not first_cell.font.bold:
                issues.append(f"'{sheet_name}' headers not bold")

        # Check auto-filter
        if not ws.auto_filter.ref:
            issues.append(f"'{sheet_name}' missing auto-filter")

    if issues:
        report.add_result(
            ValidationResult(
                level=ValidationLevel.INFO,
                check_name="formatting",
                message=f"Formatting issues: {'; '.join(issues)}",
                suggestion="Apply bold headers and auto-filter for usability",
            )
        )


# ---------------------------------------------------------------------------
# Main validation orchestrator
# ---------------------------------------------------------------------------


def validate_stm(file_path: str) -> ValidationReport:
    """Validate an STM Excel workbook."""
    report = ValidationReport(file_path)

    if not Path(file_path).exists():
        report.add_result(
            ValidationResult(
                level=ValidationLevel.CRITICAL,
                check_name="file_exists",
                message=f"File not found: {file_path}",
                suggestion="Check the file path and ensure the STM xlsx exists",
            )
        )
        return report

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        report.add_result(
            ValidationResult(
                level=ValidationLevel.CRITICAL,
                check_name="file_readable",
                message=f"Cannot open xlsx file: {e}",
                suggestion="Ensure the file is a valid .xlsx workbook",
            )
        )
        return report

    # CRITICAL checks
    check_required_sheets(wb, report)

    # Only run detailed checks if sheets exist
    sheets_present = set(wb.sheetnames)
    required_set = set(REQUIRED_SHEETS)

    if required_set.issubset(sheets_present):
        check_metadata(wb, report)
        check_source_to_bronze(wb, report)
        check_bronze_to_silver(wb, report)
        check_silver_to_gold(wb, report)

        # WARNING checks
        check_dms_traceability(wb, report)
        check_null_handling(wb, report)
        check_lineage(wb, report)
        check_code_system_mappings(wb, report)
        check_edge_cases(wb, report)
        check_column_headers(wb, report)
        check_transformation_expressions(wb, report)

        # INFO checks
        check_placeholders(wb, report)
        check_dq_references(wb, report)
        check_formatting(wb, report)
    else:
        # Still run checks on available sheets
        if "Summary" in sheets_present:
            check_metadata(wb, report)
        if "Source-to-Bronze" in sheets_present:
            check_source_to_bronze(wb, report)
        if "Bronze-to-Silver" in sheets_present:
            check_bronze_to_silver(wb, report)
        if "Silver-to-Gold" in sheets_present:
            check_silver_to_gold(wb, report)

    return report


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main():
    parser = argparse.ArgumentParser(description="Validate STM xlsx workbooks")
    parser.add_argument("path", help="STM xlsx file or directory (with --all)")
    parser.add_argument("--all", action="store_true", help="Validate all .xlsx in directory")
    parser.add_argument("--format", choices=["text", "json"], default="text")
    args = parser.parse_args()

    reports: list[ValidationReport] = []

    if args.all:
        search_dir = Path(args.path)
        xlsx_files = sorted(search_dir.rglob("*.xlsx"))
        xlsx_files = [f for f in xlsx_files if not f.name.startswith("~$")]
        if not xlsx_files:
            print(f"No .xlsx files found in {search_dir}")
            sys.exit(0)
        for f in xlsx_files:
            reports.append(validate_stm(str(f)))
    else:
        reports.append(validate_stm(args.path))

    # Output
    if args.format == "json":
        print(json.dumps([r.to_dict() for r in reports], indent=2))
    else:
        for r in reports:
            r.print_report()

    # Exit code
    max_critical = max(r.summary["CRITICAL"] for r in reports)
    max_warning = max(r.summary["WARNING"] for r in reports)

    if max_critical > 0:
        sys.exit(1)
    elif max_warning > 0:
        sys.exit(2)
    else:
        sys.exit(0)


if __name__ == "__main__":
    main()
