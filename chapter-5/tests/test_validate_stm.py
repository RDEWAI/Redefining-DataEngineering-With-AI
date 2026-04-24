"""Tests for the STM validator script."""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
import pytest
from openpyxl.styles import Font, PatternFill

# Add the validator's parent directory to sys.path
VALIDATOR_DIR = (
    Path(__file__).resolve().parent.parent
    / "mapping-analyst-plugin"
    / "skills"
    / "validate-stm"
    / "scripts"
)
sys.path.insert(0, str(VALIDATOR_DIR))

from validate_stm import (  # noqa: E402
    REQUIRED_SHEETS,
    ValidationLevel,
    ValidationReport,
    check_bronze_to_silver,
    check_code_system_mappings,
    check_dms_traceability,
    check_edge_cases,
    check_formatting,
    check_lineage,
    check_metadata,
    check_null_handling,
    check_placeholders,
    check_required_sheets,
    check_silver_to_gold,
    check_source_to_bronze,
    validate_stm,
)

# ---------------------------------------------------------------------------
# Fixtures — create xlsx workbooks using openpyxl
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
HEADER_FONT = Font(bold=True)


def _create_valid_stm(path: Path) -> Path:
    """Create a fully valid STM xlsx workbook."""
    wb = openpyxl.Workbook()

    # Summary
    ws = wb.active
    ws.title = "Summary"
    for row_idx, (key, val) in enumerate(
        [
            ("Version", "1.0"),
            ("Created", "2026-03-16"),
            ("Author", "Test Agent"),
            ("Status", "Draft"),
            ("DMS Ref", "DMS-2026-03-15-patient-360.md"),
            ("HLD Ref", "HLD-2026-03-14-pipeline.md"),
        ],
        1,
    ):
        ws.cell(row=row_idx, column=1, value=key)
        ws.cell(row=row_idx, column=2, value=val)

    # Source-to-Bronze
    ws2 = wb.create_sheet("Source-to-Bronze")
    h2 = [
        "source_table",
        "source_column",
        "source_type",
        "target_table",
        "target_column",
        "target_type",
        "transformation",
        "notes",
    ]
    for col, h in enumerate(h2, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    ws2.auto_filter.ref = f"A1:{chr(64 + len(h2))}1"
    for col, v in enumerate(
        [
            "synthea.patients",
            "id",
            "VARCHAR",
            "brz_patients",
            "id",
            "VARCHAR",
            "PASS_THROUGH",
            "PK",
        ],
        1,
    ):
        ws2.cell(row=2, column=col, value=v)

    # Bronze-to-Silver
    ws3 = wb.create_sheet("Bronze-to-Silver")
    h3 = [
        "source_table",
        "source_column",
        "source_type",
        "target_table",
        "target_column",
        "target_type",
        "transformation",
        "null_handling",
        "default_value",
        "business_rule_ref",
        "dms_ref",
    ]
    for col, h in enumerate(h3, 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    ws3.auto_filter.ref = f"A1:{chr(64 + len(h3))}1"
    for col, v in enumerate(
        [
            "brz_patients",
            "birthdate",
            "VARCHAR",
            "slv_patients",
            "date_of_birth",
            "DATE",
            "CAST(birthdate AS DATE)",
            "REJECT",
            "",
            "BR-001",
            "DMS §3.2",
        ],
        1,
    ):
        ws3.cell(row=2, column=col, value=v)

    # Silver-to-Gold
    ws4 = wb.create_sheet("Silver-to-Gold")
    h4 = [
        "target_table",
        "target_column",
        "target_type",
        "source_expression",
        "join_logic",
        "scd_type",
        "grain",
        "dms_ref",
    ]
    for col, h in enumerate(h4, 1):
        c = ws4.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    ws4.auto_filter.ref = f"A1:{chr(64 + len(h4))}1"
    for col, v in enumerate(
        [
            "dim_patient",
            "patient_sk",
            "BIGINT",
            "SEQUENCE",
            "",
            "Type 2",
            "one row per patient version",
            "DMS §5.1",
        ],
        1,
    ):
        ws4.cell(row=2, column=col, value=v)

    # Code Systems
    ws5 = wb.create_sheet("Code Systems")
    h5 = ["code_system", "source_value_pattern", "target_value", "case_expression", "notes"]
    for col, h in enumerate(h5, 1):
        c = ws5.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
    ws5.auto_filter.ref = f"A1:{chr(64 + len(h5))}1"
    ws5.cell(row=2, column=1, value="SNOMED-CT")
    ws5.cell(row=2, column=2, value="ambulatory")
    ws5.cell(row=2, column=3, value="371883000")

    # Null Handling
    ws6 = wb.create_sheet("Null Handling")
    h6 = [
        "table",
        "column",
        "layer",
        "criticality",
        "null_rate_observed",
        "action",
        "default_value",
        "business_rule_ref",
    ]
    for col, h in enumerate(h6, 1):
        c = ws6.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
    ws6.auto_filter.ref = f"A1:{chr(64 + len(h6))}1"
    ws6.cell(row=2, column=1, value="slv_patients")
    ws6.cell(row=2, column=2, value="patient_id")
    ws6.cell(row=2, column=3, value="silver")
    ws6.cell(row=2, column=4, value="HIGH")
    ws6.cell(row=2, column=5, value="0%")
    ws6.cell(row=2, column=6, value="REJECT")

    # Edge Cases
    ws7 = wb.create_sheet("Edge Cases")
    h7 = ["category", "scenario", "affected_tables", "handling_rule", "severity", "dq_rule_ref"]
    for col, h in enumerate(h7, 1):
        c = ws7.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
    ws7.auto_filter.ref = f"A1:{chr(64 + len(h7))}1"
    ws7.cell(row=2, column=1, value="Orphan Records")
    ws7.cell(row=2, column=2, value="Encounter without matching patient")
    ws7.cell(row=2, column=3, value="fact_encounter")
    ws7.cell(row=2, column=4, value="Log and quarantine")
    ws7.cell(row=2, column=5, value="HIGH")
    ws7.cell(row=2, column=6, value="DQ-001")

    # Lineage
    ws8 = wb.create_sheet("Lineage")
    h8 = [
        "gold_table",
        "gold_column",
        "silver_expression",
        "bronze_column",
        "source_column",
        "transformation_chain",
    ]
    for col, h in enumerate(h8, 1):
        c = ws8.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
    ws8.auto_filter.ref = f"A1:{chr(64 + len(h8))}1"
    for col, v in enumerate(
        [
            "dim_patient",
            "date_of_birth",
            "CAST(birthdate AS DATE)",
            "birthdate",
            "patients.birthdate",
            "PASS_THROUGH → CAST AS DATE → SCD2",
        ],
        1,
    ):
        ws8.cell(row=2, column=col, value=v)

    wb.save(str(path))
    return path


def _create_invalid_stm(path: Path) -> Path:
    """Create an STM xlsx missing most required sheets."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.cell(row=1, column=1, value="Version")
    ws.cell(row=1, column=2, value="0.1")
    wb.save(str(path))
    return path


def _create_empty_stm(path: Path) -> Path:
    """Create an STM with all sheets but no data rows."""
    wb = openpyxl.Workbook()
    wb.active.title = "Summary"
    for name in REQUIRED_SHEETS[1:]:
        wb.create_sheet(name)
    wb.save(str(path))
    return path


def _create_placeholder_stm(path: Path) -> Path:
    """Create an STM with TBD/TODO placeholders."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"
    for row_idx, (k, v) in enumerate(
        [
            ("Version", "1.0"),
            ("Created", "2026-03-16"),
            ("Author", "Test"),
            ("Status", "Draft"),
            ("DMS Ref", "TBD"),
            ("HLD Ref", "TODO"),
        ],
        1,
    ):
        ws.cell(row=row_idx, column=1, value=k)
        ws.cell(row=row_idx, column=2, value=v)

    ws2 = wb.create_sheet("Source-to-Bronze")
    for col, h in enumerate(
        [
            "source_table",
            "source_column",
            "source_type",
            "target_table",
            "target_column",
            "target_type",
            "transformation",
            "notes",
        ],
        1,
    ):
        ws2.cell(row=1, column=col, value=h)
    ws2.cell(row=2, column=1, value="TBD")
    ws2.cell(row=2, column=7, value="TODO")

    ws3 = wb.create_sheet("Bronze-to-Silver")
    for col, h in enumerate(
        [
            "source_table",
            "source_column",
            "source_type",
            "target_table",
            "target_column",
            "target_type",
            "transformation",
            "null_handling",
            "default_value",
            "business_rule_ref",
            "dms_ref",
        ],
        1,
    ):
        ws3.cell(row=1, column=col, value=h)
    ws3.cell(row=2, column=7, value="TBD")

    ws4 = wb.create_sheet("Silver-to-Gold")
    for col, h in enumerate(
        [
            "target_table",
            "target_column",
            "target_type",
            "source_expression",
            "join_logic",
            "scd_type",
            "grain",
            "dms_ref",
        ],
        1,
    ):
        ws4.cell(row=1, column=col, value=h)
    ws4.cell(row=2, column=4, value="PLACEHOLDER")

    for name in ["Code Systems", "Null Handling", "Edge Cases", "Lineage"]:
        wb.create_sheet(name)

    wb.save(str(path))
    return path


# ---------------------------------------------------------------------------
# Pytest fixtures
# ---------------------------------------------------------------------------


@pytest.fixture()
def valid_stm_file(tmp_path: Path) -> Path:
    return _create_valid_stm(tmp_path / "valid.xlsx")


@pytest.fixture()
def invalid_stm_file(tmp_path: Path) -> Path:
    return _create_invalid_stm(tmp_path / "invalid.xlsx")


@pytest.fixture()
def empty_stm_file(tmp_path: Path) -> Path:
    return _create_empty_stm(tmp_path / "empty.xlsx")


@pytest.fixture()
def placeholder_stm_file(tmp_path: Path) -> Path:
    return _create_placeholder_stm(tmp_path / "placeholder.xlsx")


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


class TestCheckRequiredSheets:
    """All 8 STM sheets must be present."""

    def test_valid_stm_passes(self, valid_stm_file):
        wb = openpyxl.load_workbook(str(valid_stm_file))
        report = ValidationReport(str(valid_stm_file))
        check_required_sheets(wb, report)
        criticals = [r for r in report.results if r.level == ValidationLevel.CRITICAL]
        assert len(criticals) == 0

    def test_missing_sheets_detected(self, invalid_stm_file):
        wb = openpyxl.load_workbook(str(invalid_stm_file))
        report = ValidationReport(str(invalid_stm_file))
        check_required_sheets(wb, report)
        criticals = [r for r in report.results if r.level == ValidationLevel.CRITICAL]
        assert len(criticals) > 0

    def test_reports_specific_missing_sheets(self, invalid_stm_file):
        wb = openpyxl.load_workbook(str(invalid_stm_file))
        report = ValidationReport(str(invalid_stm_file))
        check_required_sheets(wb, report)
        messages = " ".join(r.message for r in report.results)
        assert "Source-to-Bronze" in messages or "Bronze-to-Silver" in messages


class TestCheckMetadata:
    """Summary sheet must have complete metadata."""

    def test_valid_metadata_passes(self, valid_stm_file):
        wb = openpyxl.load_workbook(str(valid_stm_file))
        report = ValidationReport(str(valid_stm_file))
        check_metadata(wb, report)
        criticals = [r for r in report.results if r.level == ValidationLevel.CRITICAL]
        assert len(criticals) == 0

    def test_missing_metadata_detected(self, empty_stm_file):
        wb = openpyxl.load_workbook(str(empty_stm_file))
        report = ValidationReport(str(empty_stm_file))
        check_metadata(wb, report)
        criticals = [r for r in report.results if r.level == ValidationLevel.CRITICAL]
        assert len(criticals) > 0


class TestCheckSourceToBronze:
    """Source-to-Bronze sheet must have headers and data."""

    def test_valid_passes(self, valid_stm_file):
        wb = openpyxl.load_workbook(str(valid_stm_file))
        report = ValidationReport(str(valid_stm_file))
        check_source_to_bronze(wb, report)
        criticals = [r for r in report.results if r.level == ValidationLevel.CRITICAL]
        assert len(criticals) == 0

    def test_empty_sheet_fails(self, empty_stm_file):
        wb = openpyxl.load_workbook(str(empty_stm_file))
        report = ValidationReport(str(empty_stm_file))
        check_source_to_bronze(wb, report)
        criticals = [r for r in report.results if r.level == ValidationLevel.CRITICAL]
        assert len(criticals) > 0


class TestCheckBronzeToSilver:
    """Bronze-to-Silver sheet must have headers and transformation data."""

    def test_valid_passes(self, valid_stm_file):
        wb = openpyxl.load_workbook(str(valid_stm_file))
        report = ValidationReport(str(valid_stm_file))
        check_bronze_to_silver(wb, report)
        criticals = [r for r in report.results if r.level == ValidationLevel.CRITICAL]
        assert len(criticals) == 0

    def test_empty_sheet_fails(self, empty_stm_file):
        wb = openpyxl.load_workbook(str(empty_stm_file))
        report = ValidationReport(str(empty_stm_file))
        check_bronze_to_silver(wb, report)
        criticals = [r for r in report.results if r.level == ValidationLevel.CRITICAL]
        assert len(criticals) > 0


class TestCheckSilverToGold:
    """Silver-to-Gold sheet must have headers and data."""

    def test_valid_passes(self, valid_stm_file):
        wb = openpyxl.load_workbook(str(valid_stm_file))
        report = ValidationReport(str(valid_stm_file))
        check_silver_to_gold(wb, report)
        criticals = [r for r in report.results if r.level == ValidationLevel.CRITICAL]
        assert len(criticals) == 0

    def test_empty_sheet_fails(self, empty_stm_file):
        wb = openpyxl.load_workbook(str(empty_stm_file))
        report = ValidationReport(str(empty_stm_file))
        check_silver_to_gold(wb, report)
        criticals = [r for r in report.results if r.level == ValidationLevel.CRITICAL]
        assert len(criticals) > 0


class TestCheckDmsTraceability:
    """DMS references should be populated in mapping sheets."""

    def test_valid_stm_passes(self, valid_stm_file):
        wb = openpyxl.load_workbook(str(valid_stm_file))
        report = ValidationReport(str(valid_stm_file))
        check_dms_traceability(wb, report)
        warnings = [r for r in report.results if r.level == ValidationLevel.WARNING]
        assert len(warnings) == 0

    def test_empty_refs_warned(self, placeholder_stm_file):
        wb = openpyxl.load_workbook(str(placeholder_stm_file))
        report = ValidationReport(str(placeholder_stm_file))
        check_dms_traceability(wb, report)
        warnings = [r for r in report.results if r.level == ValidationLevel.WARNING]
        assert len(warnings) > 0


class TestCheckNullHandling:
    """Null Handling sheet should have criticality values."""

    def test_valid_passes(self, valid_stm_file):
        wb = openpyxl.load_workbook(str(valid_stm_file))
        report = ValidationReport(str(valid_stm_file))
        check_null_handling(wb, report)
        warnings = [r for r in report.results if r.level == ValidationLevel.WARNING]
        assert len(warnings) == 0

    def test_empty_warned(self, empty_stm_file):
        wb = openpyxl.load_workbook(str(empty_stm_file))
        report = ValidationReport(str(empty_stm_file))
        check_null_handling(wb, report)
        warnings = [r for r in report.results if r.level == ValidationLevel.WARNING]
        assert len(warnings) > 0


class TestCheckLineage:
    """Lineage sheet should have complete lineage chains."""

    def test_valid_passes(self, valid_stm_file):
        wb = openpyxl.load_workbook(str(valid_stm_file))
        report = ValidationReport(str(valid_stm_file))
        check_lineage(wb, report)
        warnings = [r for r in report.results if r.level == ValidationLevel.WARNING]
        assert len(warnings) == 0

    def test_empty_warned(self, empty_stm_file):
        wb = openpyxl.load_workbook(str(empty_stm_file))
        report = ValidationReport(str(empty_stm_file))
        check_lineage(wb, report)
        warnings = [r for r in report.results if r.level == ValidationLevel.WARNING]
        assert len(warnings) > 0


class TestCheckCodeSystemMappings:
    """Code Systems sheet should have SNOMED/RxNorm/LOINC entries."""

    def test_valid_passes(self, valid_stm_file):
        wb = openpyxl.load_workbook(str(valid_stm_file))
        report = ValidationReport(str(valid_stm_file))
        check_code_system_mappings(wb, report)
        warnings = [r for r in report.results if r.level == ValidationLevel.WARNING]
        assert len(warnings) == 0

    def test_empty_warned(self, empty_stm_file):
        wb = openpyxl.load_workbook(str(empty_stm_file))
        report = ValidationReport(str(empty_stm_file))
        check_code_system_mappings(wb, report)
        warnings = [r for r in report.results if r.level == ValidationLevel.WARNING]
        assert len(warnings) > 0


class TestCheckEdgeCases:
    """Edge Cases sheet should have documented edge cases."""

    def test_valid_passes(self, valid_stm_file):
        wb = openpyxl.load_workbook(str(valid_stm_file))
        report = ValidationReport(str(valid_stm_file))
        check_edge_cases(wb, report)
        warnings = [r for r in report.results if r.level == ValidationLevel.WARNING]
        assert len(warnings) == 0

    def test_empty_warned(self, empty_stm_file):
        wb = openpyxl.load_workbook(str(empty_stm_file))
        report = ValidationReport(str(empty_stm_file))
        check_edge_cases(wb, report)
        warnings = [r for r in report.results if r.level == ValidationLevel.WARNING]
        assert len(warnings) > 0


class TestCheckPlaceholders:
    """Should flag TBD/TODO placeholders."""

    def test_valid_has_no_placeholders(self, valid_stm_file):
        wb = openpyxl.load_workbook(str(valid_stm_file))
        report = ValidationReport(str(valid_stm_file))
        check_placeholders(wb, report)
        infos = [r for r in report.results if r.level == ValidationLevel.INFO]
        assert len(infos) == 0

    def test_placeholder_stm_flagged(self, placeholder_stm_file):
        wb = openpyxl.load_workbook(str(placeholder_stm_file))
        report = ValidationReport(str(placeholder_stm_file))
        check_placeholders(wb, report)
        infos = [r for r in report.results if r.level == ValidationLevel.INFO]
        assert len(infos) > 0


class TestCheckFormatting:
    """Should check for bold headers and auto-filter."""

    def test_formatted_stm_passes(self, valid_stm_file):
        wb = openpyxl.load_workbook(str(valid_stm_file))
        report = ValidationReport(str(valid_stm_file))
        check_formatting(wb, report)
        infos = [r for r in report.results if r.level == ValidationLevel.INFO]
        assert len(infos) == 0

    def test_unformatted_stm_flagged(self, empty_stm_file):
        wb = openpyxl.load_workbook(str(empty_stm_file))
        report = ValidationReport(str(empty_stm_file))
        check_formatting(wb, report)
        infos = [r for r in report.results if r.level == ValidationLevel.INFO]
        assert len(infos) > 0


class TestValidateStmIntegration:
    """End-to-end validation tests."""

    def test_valid_stm_no_criticals(self, valid_stm_file):
        report = validate_stm(str(valid_stm_file))
        criticals = [r for r in report.results if r.level == ValidationLevel.CRITICAL]
        assert len(criticals) == 0

    def test_invalid_stm_has_criticals(self, invalid_stm_file):
        report = validate_stm(str(invalid_stm_file))
        criticals = [r for r in report.results if r.level == ValidationLevel.CRITICAL]
        assert len(criticals) > 0

    def test_nonexistent_file(self, tmp_path):
        report = validate_stm(str(tmp_path / "nonexistent.xlsx"))
        criticals = [r for r in report.results if r.level == ValidationLevel.CRITICAL]
        assert len(criticals) > 0

    def test_report_has_file_path(self, valid_stm_file):
        report = validate_stm(str(valid_stm_file))
        assert report.file_path == str(valid_stm_file)
