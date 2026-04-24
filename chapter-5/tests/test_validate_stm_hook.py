"""Tests for the STM PostToolUse validation hook script."""

from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path

import openpyxl

HOOK_SCRIPT = (
    Path(__file__).resolve().parent.parent
    / "mapping-analyst-plugin"
    / "scripts"
    / "validate-stm-hook.py"
)

REQUIRED_SHEETS = [
    "Summary",
    "Source-to-Bronze",
    "Bronze-to-Silver",
    "Silver-to-Gold",
    "Code Systems",
    "Null Handling",
    "Edge Cases",
    "Lineage",
]


def _create_valid_stm_xlsx(path: Path) -> Path:
    """Create a minimal valid STM xlsx for testing."""
    wb = openpyxl.Workbook()

    # Summary sheet
    ws = wb.active
    ws.title = "Summary"
    metadata = [
        ("Version", "1.0"),
        ("Created", "2026-03-16"),
        ("Author", "Test Agent"),
        ("Status", "Draft"),
        ("DMS Ref", "DMS-2026-03-15-patient-360.md"),
        ("HLD Ref", "HLD-2026-03-14-pipeline.md"),
    ]
    for row_idx, (key, val) in enumerate(metadata, 1):
        ws.cell(row=row_idx, column=1, value=key)
        ws.cell(row=row_idx, column=2, value=val)

    # Source-to-Bronze
    ws2 = wb.create_sheet("Source-to-Bronze")
    headers2 = [
        "source_table",
        "source_column",
        "source_type",
        "target_table",
        "target_column",
        "target_type",
        "transformation",
        "notes",
    ]
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=col, value=h)
    row2 = [
        "synthea.patients",
        "id",
        "VARCHAR",
        "brz_patients",
        "id",
        "VARCHAR",
        "PASS_THROUGH",
        "Primary key",
    ]
    for col, v in enumerate(row2, 1):
        ws2.cell(row=2, column=col, value=v)

    # Bronze-to-Silver
    ws3 = wb.create_sheet("Bronze-to-Silver")
    headers3 = [
        "source_table",
        "source_column",
        "source_type",
        "target_table",
        "target_column",
        "target_type",
        "transformation",
        "null_handling",
        "default_value",
        "business_rule_ref",
        "dms_ref",
    ]
    for col, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=col, value=h)
    row3 = [
        "brz_patients",
        "birthdate",
        "VARCHAR",
        "slv_patients",
        "date_of_birth",
        "DATE",
        "CAST(birthdate AS DATE)",
        "REJECT",
        "",
        "BR-001",
        "DMS §3.2",
    ]
    for col, v in enumerate(row3, 1):
        ws3.cell(row=2, column=col, value=v)

    # Silver-to-Gold
    ws4 = wb.create_sheet("Silver-to-Gold")
    headers4 = [
        "target_table",
        "target_column",
        "target_type",
        "source_expression",
        "join_logic",
        "scd_type",
        "grain",
        "dms_ref",
    ]
    for col, h in enumerate(headers4, 1):
        ws4.cell(row=1, column=col, value=h)
    row4 = [
        "dim_patient",
        "patient_sk",
        "BIGINT",
        "SEQUENCE",
        "",
        "Type 2",
        "one row per patient version",
        "DMS §5.1",
    ]
    for col, v in enumerate(row4, 1):
        ws4.cell(row=2, column=col, value=v)

    # Code Systems
    ws5 = wb.create_sheet("Code Systems")
    headers5 = ["code_system", "source_value_pattern", "target_value", "case_expression", "notes"]
    for col, h in enumerate(headers5, 1):
        ws5.cell(row=1, column=col, value=h)
    ws5.cell(row=2, column=1, value="SNOMED-CT")
    ws5.cell(row=2, column=2, value="ambulatory")
    ws5.cell(row=2, column=3, value="371883000")

    # Null Handling
    ws6 = wb.create_sheet("Null Handling")
    headers6 = [
        "table",
        "column",
        "layer",
        "criticality",
        "null_rate_observed",
        "action",
        "default_value",
        "business_rule_ref",
    ]
    for col, h in enumerate(headers6, 1):
        ws6.cell(row=1, column=col, value=h)
    ws6.cell(row=2, column=1, value="slv_patients")
    ws6.cell(row=2, column=2, value="patient_id")
    ws6.cell(row=2, column=3, value="silver")
    ws6.cell(row=2, column=4, value="HIGH")

    # Edge Cases
    ws7 = wb.create_sheet("Edge Cases")
    headers7 = [
        "category",
        "scenario",
        "affected_tables",
        "handling_rule",
        "severity",
        "dq_rule_ref",
    ]
    for col, h in enumerate(headers7, 1):
        ws7.cell(row=1, column=col, value=h)
    ws7.cell(row=2, column=1, value="Orphan Records")
    ws7.cell(row=2, column=2, value="Encounter without matching patient")

    # Lineage
    ws8 = wb.create_sheet("Lineage")
    headers8 = [
        "gold_table",
        "gold_column",
        "silver_expression",
        "bronze_column",
        "source_column",
        "transformation_chain",
    ]
    for col, h in enumerate(headers8, 1):
        ws8.cell(row=1, column=col, value=h)
    row8 = [
        "dim_patient",
        "date_of_birth",
        "CAST(birthdate AS DATE)",
        "birthdate",
        "patients.birthdate",
        "PASS_THROUGH → CAST AS DATE → SCD2",
    ]
    for col, v in enumerate(row8, 1):
        ws8.cell(row=2, column=col, value=v)

    wb.save(str(path))
    return path


def _run_hook(stdin_data: str) -> subprocess.CompletedProcess:
    """Run the hook script with the given stdin."""
    return subprocess.run(
        [sys.executable, str(HOOK_SCRIPT)],
        input=stdin_data,
        capture_output=True,
        text=True,
        timeout=30,
    )


def _write_input(file_path: str) -> str:
    """Create a PostToolUse JSON input for the hook."""
    return json.dumps({"tool_input": {"file_path": file_path}})


class TestHookSkipsNonStmFiles:
    """Hook should exit 0 for files that are not STM documents."""

    def test_non_xlsx_file(self):
        result = _run_hook(_write_input("/project/outputs/stm/data.md"))
        assert result.returncode == 0

    def test_xlsx_not_in_outputs_stm(self):
        result = _run_hook(_write_input("/project/some/other/file.xlsx"))
        assert result.returncode == 0

    def test_gitkeep_file(self):
        result = _run_hook(_write_input("/project/outputs/stm/.gitkeep"))
        assert result.returncode == 0

    def test_empty_file_path(self):
        result = _run_hook(json.dumps({"tool_input": {"file_path": ""}}))
        assert result.returncode == 0

    def test_missing_tool_input(self):
        result = _run_hook(json.dumps({"something": "else"}))
        assert result.returncode == 0


class TestHookValidatesStmFiles:
    """Hook should validate .xlsx files in outputs/stm/."""

    def test_valid_stm_passes(self, tmp_path):
        stm_dir = tmp_path / "outputs" / "stm"
        stm_dir.mkdir(parents=True)
        stm_file = stm_dir / "test.xlsx"
        _create_valid_stm_xlsx(stm_file)
        result = _run_hook(_write_input(str(stm_file)))
        assert result.returncode == 0

    def test_critical_issues_block(self, tmp_path):
        stm_dir = tmp_path / "outputs" / "stm"
        stm_dir.mkdir(parents=True)
        stm_file = stm_dir / "bad.xlsx"
        # Create xlsx with only one sheet (missing required sheets)
        wb = openpyxl.Workbook()
        wb.active.title = "Summary"
        wb.save(str(stm_file))
        result = _run_hook(_write_input(str(stm_file)))
        assert result.returncode == 2

    def test_empty_stm_produces_critical(self, tmp_path):
        stm_dir = tmp_path / "outputs" / "stm"
        stm_dir.mkdir(parents=True)
        stm_file = stm_dir / "empty.xlsx"
        wb = openpyxl.Workbook()
        wb.active.title = "Sheet1"
        wb.save(str(stm_file))
        result = _run_hook(_write_input(str(stm_file)))
        assert result.returncode == 2


class TestHookInvalidInput:
    """Hook should handle invalid JSON gracefully."""

    def test_invalid_json_input(self):
        result = _run_hook("not json at all")
        assert result.returncode == 0

    def test_empty_stdin(self):
        result = _run_hook("")
        assert result.returncode == 0
